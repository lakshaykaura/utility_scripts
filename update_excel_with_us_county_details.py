import pandas as pd
from bs4 import BeautifulSoup
import requests
import os
import logging
from openpyxl import load_workbook
from rich.console import Console
from rich.logging import RichHandler
import utils.excel as excel_utils

console = Console()

logging.basicConfig(
    level="INFO",
    format="%(message)s",
    datefmt="[%X]",
    handlers=[RichHandler(rich_tracebacks=True, console=console)],
)

EMPTY_STRING = ""

excel_file = r"C:\Users\lkaura\Desktop\Scripts\resources\ZipCodeCountyMapper.xlsx"

backup_file = excel_utils.backup_excel_file(excel_file)
console.print("[yellow]Backup of the Excel file created successfully.[/yellow]")

sheet_name = "Sheet1"


def clear_console():
    """Clear the console screen."""
    os.system("cls" if os.name == "nt" else "clear")


def print_progress_bar(
    iteration, total, prefix="", suffix="", length=50, fill="█", print_end="\r"
):
    """Call in a loop to create terminal progress bar"""
    percent = ("{0:.2f}").format(100 * (iteration / float(total)))
    filled_length = int(length * iteration // total)
    bar = fill * filled_length + "-" * (length - filled_length)
    print(f"\r{prefix} |{bar}| {percent}% {suffix}", end=print_end)
    # Print New Line on Complete
    if iteration == total:
        print()


def fetch_details(zip_code):
    zip_code_str = str(zip_code).zfill(5)
    url = f"https://www.geonames.org/postalcode-search.html?q={zip_code_str}&country=US"
    try:
        response = requests.get(url)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, "html.parser")
            table = soup.find("table", class_="restable")
            if table:
                for row in table.find_all("tr")[1:]:
                    cells = row.find_all("td")
                    if len(cells) > 2 and zip_code_str in cells[2].text.strip():
                        city = cells[1].text.strip()
                        state = cells[4].text.strip()
                        county = cells[5].text.strip()
                        if county.endswith(" (city)"):
                            county = county.replace(" (city)", "")
                        return county, city, state
            return EMPTY_STRING, EMPTY_STRING, EMPTY_STRING
    except requests.exceptions.RequestException as e:
        print(f"Request exception for ZIP code {zip_code_str}: {e}")
        return EMPTY_STRING, EMPTY_STRING, EMPTY_STRING


zip_codes_df = pd.read_excel(excel_file, sheet_name=sheet_name)

# Count rows that need processing
needs_processing = zip_codes_df.apply(
    lambda row: pd.isnull(row["County"])
    or pd.isnull(row["City"])
    or pd.isnull(row["State"])
    or row["County"] == EMPTY_STRING
    or row["City"] == EMPTY_STRING
    or row["State"] == EMPTY_STRING,
    axis=1,
)
total_rows = needs_processing.sum()
clear_console()

wb = load_workbook(excel_file)
ws = wb[sheet_name]

processed_rows = 0  # Counter for rows actually processed

for index, row in zip_codes_df.iterrows():
    if needs_processing[index]:  # Check if this row needs processing
        zip_code = row["ZipCode"]
        county, city, state = fetch_details(zip_code)

        ws[f"B{index + 2}"] = county
        ws[f"C{index + 2}"] = city
        ws[f"D{index + 2}"] = state

        processed_rows += 1  # Increment the counter only when a row is processed
        print_progress_bar(
            processed_rows, total_rows, prefix="Progress:", suffix="Complete", length=50
        )

wb.save(excel_file)
wb.close()
print("All ZIP codes processed and updated successfully.")
