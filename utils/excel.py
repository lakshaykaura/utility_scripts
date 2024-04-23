import os
import xlwings as xw
import shutil
import win32com.client
import openpyxl
import xlrd

def run_vba_macro_on_files_in_directory(directory_path, macro_name):
    """
    Runs a specified VBA macro on all Excel files in a given directory.

    Args:
        directory_path (str): The path to the directory containing the Excel files.
        macro_name (str): The name of the VBA macro to run.

    This function opens each Excel file in the specified directory and runs the specified VBA macro on it.
    The Excel application is kept in the background during this process. The function ensures that the
    PERSONAL.XLSB workbook is open so that the macro can be accessed. After running the macro, each workbook
    is saved and closed. Finally, the Excel application is quit.
    """
    # Create an Excel application object
    excel_app = win32com.client.Dispatch("Excel.Application")
    excel_app.Visible = False  # We keep Excel in the background

    # Ensure PERSONAL.XLSB is open to access the macro
    try:
        personal_wb = excel_app.Workbooks("PERSONAL.XLSB")
    except:
        personal_wb = excel_app.Workbooks.Open(
            excel_app.StartupPath + r"\PERSONAL.XLSB"
        )

    # Loop through all files in the directory
    for filename in os.listdir(directory_path):
        if filename.endswith(".xls") or filename.endswith(".xlsx"):
            file_path = os.path.join(directory_path, filename)

            # Open the workbook
            wb = excel_app.Workbooks.Open(file_path)

            # Run the macro
            excel_app.Run(macro_name)

            # Close the workbook
            wb.Close(SaveChanges=True)

    # Quit the Excel application
    excel_app.Quit()


def backup_excel_file(file_path):
    """
    Creates a backup of the Excel file.

    Args:
        file_path (str): The full path to the Excel file to be backed up.

    Returns:
        str: The file path of the backed-up Excel file.
    """
    backup_file_path = file_path.replace(".xlsx", "_backup.xlsx")
    shutil.copy(file_path, backup_file_path)
    return backup_file_path


def open_workbook(file_path, visible=False):
    """
    Opens an Excel workbook.

    Args:
        file_path (str): The full path to the Excel file to be opened.
        visible (bool): Specifies whether the Excel application should be visible.

    Returns:
        tuple: A tuple containing the Excel app and workbook instances.
    """
    app = xw.App(visible=visible)
    wb = app.books.open(file_path)
    return app, wb


def get_column_letter_from_index(index):
    """
    Converts a 1-based column index into its corresponding Excel column letter.

    Excel columns are labeled alphabetically starting with 'A' for the first column, 'B' for the second, and so on,
    up to 'Z' for the 26th column. After 'Z', columns continue with double letters starting from 'AA', 'AB', etc.
    This function calculates the corresponding Excel column letter for a given 1-based column index.

    Parameters:
    - index (int): The 1-based index of the column. The first column has an index of 1, not 0.

    Returns:
    - str: The Excel column letter corresponding to the provided index.

    Raises:
    - ValueError: If the provided index is less than 1 since Excel columns are 1-based.

    Note: This function is designed to work with indices that represent valid Excel column positions. Excel supports
    up to 16,384 columns as of Excel 2019 and Office 365, corresponding to column 'XFD'.
    """
    if index < 1:
        raise ValueError("Index is too small")
    result = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def get_column_letter_by_header(sheet, header_row, last_col):
    """
    Creates a dictionary mapping header names to column letters in the specified header row.

    :param sheet: The xlwings Sheet object.
    :param header_row: The row number of the header row.
    :return: A dictionary mapping header names to column letters.
    """
    headers = sheet.range(f"{header_row}:{header_row}").value[:last_col]
    return {
        header: get_column_letter_from_index(index + 1)
        for index, header in enumerate(headers)
        if header is not None
    }


def get_jira_ids_from_excel_sheet(sheet, last_row, column_map):
    """
    Extracts Jira IDs from an Excel sheet and maps them to their row numbers.

    Args:
        sheet (xlwings.Sheet): The Excel sheet to extract Jira IDs from.
        last_row (int): The last row in the sheet to consider.
        column_map (dict): A dictionary mapping column names to their letter representations.

    Returns:
        dict: A dictionary mapping Jira IDs to their row numbers.

    This function iterates over the rows of the specified Excel sheet, extracting the Jira ID from column 'A'
    and the status from the column specified in the column map. If the Jira ID is not None and the status is not 'Done',
    the function adds the Jira ID to the dictionary with its row number as the value.
    """
    status_column = column_map["Status"]
    jira_id_to_row_map = {}

    for row in range(2, last_row + 1):
        jira_id_cell = f"A{row}"
        status_cell = f"{status_column}{row}" if status_column else None

        jira_id = sheet.range(jira_id_cell).value
        status = sheet.range(status_cell).value if status_cell else None

        if jira_id and status != "Done":
            jira_id_to_row_map[jira_id] = row

    return jira_id_to_row_map


def search_xlsx(file_name, keyword):
    """
    Searches for a keyword in an Excel file.

    Args:
        file_name (str): The name of the Excel file to search.
        keyword (str): The keyword to search for.

    Returns:
        bool: True if the keyword is found, False otherwise.

    This function loads the specified Excel file and iterates over each sheet, row, and cell in the workbook.
    If a cell contains a value and the keyword is found within that value, the function returns True.
    If the keyword is not found in any cell in the workbook, the function returns False.
    """
    workbook = openpyxl.load_workbook(file_name)
    for sheet in workbook:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and keyword in str(cell.value):
                    return True
    return False


def search_xls(file_name, keyword):
    """
    Searches for a keyword in an Excel (.xls) file.

    Args:
        file_name (str): The name of the Excel file to search.
        keyword (str): The keyword to search for.

    Returns:
        bool: True if the keyword is found, False otherwise.

    This function loads the specified Excel file using the xlrd library and iterates over each sheet, row, and cell in the workbook.
    If a cell contains a value and the keyword is found within that value, the function returns True.
    If the keyword is not found in any cell in the workbook, the function returns False.
    """
    workbook = xlrd.open_workbook(file_name)
    for sheet in workbook.sheets():
        for row in range(sheet.nrows):
            for col in range(sheet.ncols):
                cell = sheet.cell(row, col)
                if cell.value and keyword in str(cell.value):
                    return True
    return False


def search_for_keyword_in_excel_files_within_directory(directory, keyword):
    """
    Searches for a keyword in all Excel files within a specified directory.

    Args:
        directory (str): The directory to search within.
        keyword (str): The keyword to search for.

    This function walks through the specified directory and its subdirectories, checking each file.
    If a file is an Excel file (either .xls or .xlsx), the function searches for the keyword within the file.
    If the keyword is found, the function opens the file using the default application.
    """
    for root, _, files in os.walk(directory):
        for file in files:
            if file.endswith((".xls", ".xlsx")):
                file_path = os.path.join(root, file)
                print(f"\nChecking file {file_path} for keyword {keyword}...\n")
                if file.endswith(".xlsx"):
                    found = search_xlsx(file_path, keyword)
                else:
                    found = search_xls(file_path, keyword)
                if found:
                    os.startfile(file_path)
